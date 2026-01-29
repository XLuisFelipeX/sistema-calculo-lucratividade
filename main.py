from openpyxl import load_workbook

CLIENTES_ELEGIVEIS_GV8 = {
    "EMPRESA XYZ LTDA",
}

TARIFAS_VALIDAS = [
    "CUSTO REGISTRO BOLETO ONLINE",
    "CUSTO ENVIO TED",
    "CUSTO ENVIO PIX",
    "CUSTO RECEBIMENTO PIX",
    "SPLIT PERCENTUAL",
    "MANUTENÇÃO DE CONTA"
]

def eh_tarifa_valida(descricao):
    """
    Retorna True se a descrição da operação
    for uma tarifa válida segundo a whitelist.
    """
    return descricao in TARIFAS_VALIDAS

def ler_planilha_privilege(caminho_arquivo, sistema_origem):
    workbook = load_workbook(caminho_arquivo)
    sheet = workbook.active

    linhas = list(sheet.iter_rows(values_only=True))

    cabecalho = linhas[0]
    dados = linhas[1:]  # ignora o cabeçalho

    registros = []

    for linha in dados:
        registro = {
            "data": linha[6],        # Coluna G - DataMovto
            "descricao": linha[7],   # Coluna H - Operação
            "valor": linha[8],       # Coluna I - Valor
            "sistema_origem": sistema_origem,
            "cliente": "EMPRESA XYZ LTDA",
            "cliente_elegivel": True
        }
        registros.append(registro)

    return registros

def aplicar_elegibilidade_por_sistema(registro):
    sistema = registro["sistema_origem"]

    if sistema in ["FOURBANK", "AARIN"]:
        registro["cliente_elegivel"] = True
    elif sistema == "PRIVILEGE":
        cliente = registro["cliente"]
        registro["cliente_elegivel"] = cliente in CLIENTES_ELEGIVEIS_GV8
    else:
        registro["cliente_elegivel"] = False

def filtrar_tarifas(registros):
    """
    Recebe uma lista de registros e retorna
    apenas aqueles que são tarifas válidas.
    """
    tarifas = []

    for registro in registros:
        descricao = registro["descricao"]

        if eh_tarifa_valida(descricao):
            tarifas.append(registro)

    return tarifas

def classificar_operacoes(registros):
    """
    Classifica cada registro como 'tarifa' ou 'neutra'
    com base na regra de whitelist.
    """
    for registro in registros:
        descricao = registro["descricao"]

        if eh_tarifa_valida(descricao):
            registro["categoria_operacao"] = "tarifa"
        else:
            registro["categoria_operacao"] = "neutra"

    return registros

def classificar_impacto_lucro(registro):
    """
    Atribui o impacto financeiro do registro.
    Não altera elegibilidade, categoria ou valor.
    """
    if registro.get("categoria_operacao") != "tarifa":
        registro["impacto_lucro"] = "neutro"
    else:
        registro["impacto_lucro"] = "receita"

def filtrar_registros_elegiveis(registros):
    """
    Retorna apenas registros com cliente_elegivel = True.
    Não altera os registros originais.
    """
    registros_elegiveis = []

    for registro in registros:
        if registro.get("cliente_elegivel") is True:
            registros_elegiveis.append(registro)

    return registros_elegiveis

def somar_valor_tarifas(registros):
    """
    Soma o valor de todos os registros classificados como tarifa.
    """
    total = 0

    for registro in registros:
        if registro["categoria_operacao"] == "tarifa":
            total += registro["valor"]

    return total


def main():
    caminho_arquivo = r"C:\Users\Luis.Nunes\Desktop\Projeto Sistema de Cálculo de Lucratividade\Relatório Privilege - 01.12-31.12\Relatório Movimentos - 01.12-31.12.xlsx"

    registros = ler_planilha_privilege(caminho_arquivo, sistema_origem="PRIVILEGE")

    for registro in registros:
        aplicar_elegibilidade_por_sistema(registro)

    registros_classificados = classificar_operacoes(registros)

    for registro in registros_classificados:
        classificar_impacto_lucro(registro)

    # ==================================================
    # TESTE MANUAL – SPRINT 4.3
    # Validação do impacto financeiro por registro
    # ==================================================

    # for r in registros_classificados[:5]:
    #     print(
    #         r["descricao"],
    #         r["categoria_operacao"],
    #         r["impacto_lucro"],
    #         r["cliente_elegivel"]
    #     )

    # ==================================================
    # FIM DO TESTE MANUAL – SPRINT 4.3
    # ==================================================

    registros_elegiveis = filtrar_registros_elegiveis(registros_classificados)

    total_tarifas = somar_valor_tarifas(registros_elegiveis)
    print(f"Total de tarifas elegíveis: R$ {total_tarifas:.2f}")

    # ==================================================
    # TESTES MANUAIS HISTÓRICOS (SPRINTS ANTERIORES)
    # Mantidos apenas para consulta
    # ==================================================

    # print(registros[0])

    # for r in registros[:5]:
    #     print(r["cliente"], r["sistema_origem"], r["cliente_elegivel"])

    # ==================================================
    # FIM DOS TESTES MANUAIS HISTÓRICOS
    # ==================================================

    # ==================================================
    # TESTE MANUAL – VALIDAÇÃO DA CLASSIFICAÇÃO
    # ==================================================

    # print("\nTeste de classificação (3 primeiros registros):")
    # for r in registros_classificados[:3]:
    #    print(r)

    # ==================================================
    # FIM DO TESTE DE CLASSIFICAÇÃO
    # ==================================================

    # ==================================================
    # TESTE MANUAL – VALIDAÇÃO DA SOMA DE TARIFAS
    # ==================================================
   
    # total_tarifas = somar_valor_tarifas(registros_classificados)

    # print(f"\nTotal de tarifas no período: R$ {total_tarifas:.2f}")

    # ==================================================
    # FIM DO TESTE DE SOMA DE TARIFAS
    # ==================================================

    # ==================================================
    # TESTE MANUAL – VALIDAÇÃO DE VOLUME DE DADOS
    # Usado para confirmar leitura completa da planilha
    # e aplicação correta da regra de tarifas.
    # ==================================================

    # print(f"Total de registros lidos: {len(registros)}")
    # print(f"Total de tarifas válidas: {len(tarifas)}")
    # OBS: 'tarifas' era usado antes da classificação.
    # Hoje a fonte de verdade é 'categoria_operacao'.

    # ==================================================
    # FIM DO TESTE DE VOLUME
    # ==================================================


    # ==================================================
    # TESTE MANUAL – VALIDAÇÃO DA REGRA DE TARIFA
    # Usado apenas para aprendizado e verificação pontual.
    # Pode ser reativado se necessário.
    # ==================================================

    # tarifas = filtrar_tarifas(registros)

    # print("\nTeste da regra de tarifa:")

    # primeiro_registro = registros[0]
    # descricao = primeiro_registro["descricao"]

    # print("Descrição:", descricao)
    # print("É tarifa válida?", eh_tarifa_valida(descricao))

    # ==================================================
    # FIM DO TESTE MANUAL
    # ==================================================


    # ==================================================
    # TESTE MANUAL – VERIFICAÇÃO DA LEITURA DA PLANILHA
    # ==================================================

    # print("Primeiros 5 registros lidos:")
    # for r in registros[:5]:
    #     print(r)

    # ==================================================
    # FIM DO TESTE DE LEITURA
    # ==================================================

    # ==================================================
    # SPRINT 5 — ENCARGOS OPERACIONAIS (ESTRUTURA)
    # ==================================================

    # Tarifas que NÃO possuem encargo operacional
    # - SPLIT PERCENTUAL → Cash In
    # - MANUTENÇÃO DE CONTA → Mensalidade
    TARIFAS_ISENTAS_DE_ENCARGO = {
        "SPLIT PERCENTUAL",
        "MANUTENÇÃO DE CONTA",
    }

    # Definição dos encargos operacionais externos à planilha
    # Estes encargos representam custos do banco para o GV8
    ENCARGOS_OPERACIONAIS = [
        {
            "id_encargo": "ENCARGO_BMP_TARIFAS",
            "descricao": "Encargo operacional cobrado pelo banco BMP sobre tarifas",
            "tipo": "percentual",
            "percentual": 0.02,  # 2%
            "banco_liquidante": "BMP",
            "ativo": True,
        }
    ]

    # ==================================================
    # FIM — SPRINT 5 (MICRO-PASSO 5.8)
    # ==================================================

    # ==================================================
    # SPRINT 5 — APLICAÇÃO DE ENCARGOS OPERACIONAIS
    # ==================================================

    def aplicar_encargos(registros_elegiveis):
        """
        Aplica encargos operacionais sobre registros elegíveis,
        sem alterar os registros originais.

        Retorna uma lista de encargos aplicados ou não aplicados,
        sempre com motivo explícito.
        """

        encargos_aplicados = []

        for registro in registros_elegiveis:
            descricao = registro.get("descricao")
            valor = registro.get("valor")
            sistema_origem = registro.get("sistema_origem")
            categoria = registro.get("categoria_operacao")
            impacto = registro.get("impacto_lucro")

            # =============================
            # Regras de bloqueio
            # =============================

            if registro.get("cliente_elegivel") is not True:
                encargos_aplicados.append({
                    "registro": registro,
                    "encargo_aplicado": False,
                    "motivo_aplicacao": "Registro não elegível ao GV8"
                })
                continue

            if categoria != "tarifa":
                encargos_aplicados.append({
                    "registro": registro,
                    "encargo_aplicado": False,
                    "motivo_aplicacao": "Operação não é tarifa"
                })
                continue

            if impacto != "receita":
                encargos_aplicados.append({
                    "registro": registro,
                    "encargo_aplicado": False,
                    "motivo_aplicacao": "Operação sem impacto de receita"
                })
                continue

            if descricao in TARIFAS_ISENTAS_DE_ENCARGO:
                motivo = (
                    "Tarifa do tipo Cash In isenta de encargo operacional"
                    if descricao == "SPLIT PERCENTUAL"
                    else "Tarifa do tipo Mensalidade isenta de encargo operacional"
                )

                encargos_aplicados.append({
                    "registro": registro,
                    "encargo_aplicado": False,
                    "motivo_aplicacao": motivo
                })
                continue

            # =============================
            # Aplicação do encargo
            # =============================

            encargo_encontrado = False

            for encargo in ENCARGOS_OPERACIONAIS:
                if not encargo.get("ativo"):
                    continue

                if sistema_origem != "PRIVILEGE":
                    continue

                percentual = encargo.get("percentual")

                if percentual is None or percentual <= 0 or percentual > 1:
                    encargos_aplicados.append({
                        "registro": registro,
                        "encargo_aplicado": False,
                        "motivo_aplicacao": "Percentual de encargo inválido"
                    })
                    encargo_encontrado = True
                    break

                if valor is None or valor <= 0:
                    encargos_aplicados.append({
                        "registro": registro,
                        "encargo_aplicado": False,
                        "motivo_aplicacao": "Valor da tarifa inválido para aplicação de encargo"
                    })
                    encargo_encontrado = True
                    break

                valor_encargo = valor * percentual

                encargos_aplicados.append({
                    "registro": registro,
                    "encargo_aplicado": True,
                    "id_encargo": encargo.get("id_encargo"),
                    "percentual_aplicado": percentual,
                    "valor_base_tarifa": valor,
                    "valor_encargo_calculado": round(valor_encargo, 2),
                    "motivo_aplicacao": "Tarifa operacional elegível ao encargo BMP"
                })

                encargo_encontrado = True
                break

            if not encargo_encontrado:
                encargos_aplicados.append({
                    "registro": registro,
                    "encargo_aplicado": False,
                    "motivo_aplicacao": "Nenhum encargo aplicável ao registro"
                })

        return encargos_aplicados

    # ==================================================
    # FIM — SPRINT 5 (MICRO-PASSO 5.9)
    # ==================================================

    # ==================================================
    # SPRINT 5 — VALIDAÇÃO DA APLICAÇÃO DE ENCARGOS
    # (somente print, sem impacto em cálculo)
    # ==================================================

    encargos_aplicados = aplicar_encargos(registros_elegiveis)

    print("\n--- Validação Sprint 5 | Encargos Operacionais ---")

    for item in encargos_aplicados[:10]:  # limita para não poluir o console
        registro = item["registro"]

        if item["encargo_aplicado"]:
            print(
                f"[ENCARGO APLICADO] "
                f"Operação: {registro['descricao']} | "
                f"Valor Tarifa: R$ {registro['valor']:.2f} | "
                f"Encargo: R$ {item['valor_encargo_calculado']:.2f} | "
                f"Motivo: {item['motivo_aplicacao']}"
            )
        else:
            print(
                f"[SEM ENCARGO] "
                f"Operação: {registro['descricao']} | "
                f"Motivo: {item['motivo_aplicacao']}"
            )

    print("--- Fim da validação Sprint 5 ---\n")

if __name__ == "__main__":
    main()